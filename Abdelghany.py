from pyomo.environ import *
import openpyxl as x
import pyomo.opt as pyo
def optimize(stg_dat):    
    model = ConcreteModel()

    # Sets rf import
    acwb = x.load_workbook("aircraft.xlsx")
    acws = acwb.active
    def R_init(model):
       return [acws[i][0].value for i in range(2, acws.max_row)]
    model.R = Set(initialize=R_init)
    #model.R.pprint()
    def F_init(model):
        return [stg_dat[i].flight_id for i in range(len(stg_dat))]
    model.F = Set(initialize=F_init)
    #model.F.pprint()

    def c_init(model, r, f):
        for i in range(2, acws.max_row):
            for j in range(len(stg_dat)):
                if ((acws[i][8].value==stg_dat[j].origin and (acws[i][0].value==r)) and stg_dat[j].flight_id==f):
                    print(i,j,stg_dat[j].origin) 
                    return acws[i][5].value*(stg_dat[j].planned_arrival-stg_dat[j].planned_departure)/60
                else:
                    return 999999
    model.c = Param(model.R, model.F, initialize=c_init)
    model.c.pprint()
    model.cd = Param(model.F, default=(0.75*120))
    #model.cd.pprint()
    def t_init(model, f):
        for i in range(len(stg_dat)):
            if stg_dat[i].flight_id == f:
                return stg_dat[i].planned_departure
            else:
                return 99999
    model.t = Param(model.F, initialize=t_init)
    #model.t.pprint()     
    model.cc = Param(model.F, default=5000)
    #model.cc.pprint()
    model.b = Param(model.R, model.F, default=1)
    #model.b.pprint
    def a_init(model, r, f):
        for i in range(2, acws.max_row):
            for j in range(len(stg_dat)):
                if (acws[i][8].value==stg_dat[j].origin and acws[i][0].value==r and stg_dat[j].flight_id==f):
                    return 0.5
                else:
                    return 99999
    model.a = Param(model.R, model.F, initialize=a_init)
    #model.a.pprint()
    def T_init(model, f):
        for j in range(len(stg_dat)):
            if stg_dat[j].flight_id==f :
                return (stg_dat[j].planned_arrival-stg_dat[j].planned_departure)
            else:
                return 99999
    model.T = Param(model.F, initialize=T_init)
    #model.T.pprint()
    # Variables
    model.x = Var((model.R,model.F), domain=Binary)
    model.L = Var(model.F, domain=Binary)
    model.m = Var(model.F, domain=NonNegativeIntegers)
    model.n = Var(model.F, domain=NonNegativeIntegers)


    def obj_expression(model):
        return (sum( model.c[r, f] * model.x[r,f] for r in model.R for f in model.F ) 
            +sum(model.cd[f] * (model.m[f]-model.t[f]) for f in model.F)
            +sum(model.cc[f] * model.L[f] for f in model.F))

    model.obj = Objective(rule=obj_expression,sense = minimize)


    '''def C1(model, r, f):
        return model.x[r,f] <= model.b[r,f]

    def C2(model, r):
        return sum(model.x[r, f] for f in model.F) <= 1

    def C3(model, f):
        return sum(model.x[r, f] for r in model.R) == 1 - model.L[f]

    def C4(model,r, f):
        return model.m[f] >= model.x[r,f] * model.a[r,f]

    def C5(model, f):
        return model.n[f] == model.m[f] + model.T[f]

    def C6(model, f):
        return model.m[f] >= model.t[f]



    model.Co1 = Constraint((model.R,model.F), rule=C1)
    model.Co2 = Constraint(model.R, rule=C2)
    model.Co3 = Constraint(model.F, rule=C3)
    model.Co4 = Constraint((model.R,model.F), rule=C4)
    model.Co5 = Constraint(model.F, rule=C5)
    model.Co6 = Constraint(model.F, rule=C6)
    '''

    

    
    #instance = model.create_instance()
    opt = pyo.SolverFactory('cplex')
    opt.solve(model,tee=True)
    model.display()
    return 0
    #
