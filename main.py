import openpyxl as x
from pyomo.environ import *
wb = x.load_workbook("flights.xlsx")
ws = wb.active
acwb = x.load_workbook("aircraft.xlsx")
acws = acwb.active
apwb = x.load_workbook("airports.xlsx")
apws = apwb.active
from Staging import staging
from Abdelghany import optimize


current_stage=1


while current_stage <= 24:
    
    
    print('current stage=',current_stage)
    st = staging(current_stage,ws)
    for i in range(len(st)):
            for row in range(2, ws.max_row+1):
                if (ws[row][0].value == st[i].flight_id):
                    ws[row][4].value = current_stage
    
    op=optimize(st,acws,apws,current_stage)

    
    for f in op.F:
        k = 0
        for r in op.R:
            for i in op.I:
                for j in op.I:
                    tmp = int(value(op.x[r,f,i,j]) + 0.5)
                    k += tmp
                    if tmp:
                        for row1 in range(2, ws.max_row+1):
                            if  ws[row1][0].value == f:
                                ws[row1][7].value = value(op.m[f])
                                ws[row1][8].value = value(op.n[f])
                                for row2 in range(2, acws.max_row+1):
                                    if  acws[row2][0].value == r:
                                        acws[row2][7].value = value(op.n[f])
                                        acws[row2][8].value = ws[row1][6].value
            
                                
                                
        for row in range(2, ws.max_row+1):
            if  int(ws[row][0].value) == f:
                ws[row][9].value = k
                L = int(value(op.L[f]))
                ws[row][10].value = L
                if k + L != 1:
                    print(f'Error, Flight {f} has the problem!')
                    input('Continue?')                    

    
    
                              


    current_stage+=1
    
    wb.save("flights.xlsx")
    acwb.save("aircraft.xlsx")



