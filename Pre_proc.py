import openpyxl as x
ws = x.load_workbook("flights.xlsx").active


current_stage=1

class flight:
    def __init__(self,flight_id,planned_departure,planned_arrival,included_in_stage,delay):
        self.flight_id=flight_id
        self.planned_departure=planned_departure
        self.planned_arrival=planned_arrival
        self.included_in_stage=included_in_stage
        self.delay=delay

f_list=[]
for row in range(2, ws.max_row):
    f=flight(ws[row][0].value,ws[row][1].value,ws[row][2].value,0,ws[row][3].value)
    f_list.append(f)


# step 1 : sorting flights chronologicaly (based on departure times)
f_sorted=sorted(f_list,key=lambda x: x.planned_departure)



# step 2 : finding the critical flights (disrupted and not in previous stages)
critical_flight = None
critical_flight_arrival = None
for i in range(len(f_sorted)):
    if(f_sorted[i].delay > 0 and f_sorted[i].included_in_stage==0 ):
        critical_flight=f_sorted[i].flight_id
        f_sorted[i].included_in_stage=current_stage
        critical_flight_arrival = f_sorted[i].planned_arrival
        break

# step 3 : listing current stage flights (not in prevoius stages and depart before the arrival of critical flight)
current_stage_flights=[critical_flight]
for i in range(len(f_sorted)):
    if(f_sorted[i].included_in_stage==0 and f_sorted[i].planned_departure < critical_flight_arrival):
        current_stage_flights.append(f_sorted[i].flight_id)
        f_sorted[i].included_in_stage=current_stage


#print results
print(f'Current stage flights(stage:{current_stage}): {current_stage_flights}')

'''for i in range(len(f_sorted)):
    print(f_sorted[i].flight_id),
    print(f_sorted[i].planned_departure),
    print(f_sorted[i].planned_arrival),
    print(f_sorted[i].included_in_stage),
    print(f_sorted[i].delay,'\n')'''
