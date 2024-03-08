import openpyxl as x
def staging():
    wb = x.load_workbook("flights.xlsx")
    ws = wb.active


    current_stage=1


    class flight:
        def __init__(self,flight_id,planned_departure,planned_arrival,included_in_stage,delay,origin,dest):
            self.flight_id=flight_id
            self.planned_departure=planned_departure
            self.planned_arrival=planned_arrival
            self.included_in_stage=included_in_stage
            self.delay=delay
            self.origin=origin
            self.dest=dest
    f_list=[]
    for row in range(2, ws.max_row):
        f=flight(ws[row][0].value,ws[row][1].value,ws[row][2].value,ws[row][4].value,ws[row][3].value,ws[row][5].value,ws[row][6].value)
        f_list.append(f)


    # step 1 : sorting flights chronologicaly (based on departure times)
    f_sorted=sorted(f_list,key=lambda x: x.planned_departure)



    # step 2 : finding the critical flights (disrupted and not in previous stages)
    current_stage_flights=[]
    for i in range(len(f_sorted)):
        if(f_sorted[i].delay > 0 and f_sorted[i].included_in_stage==0 ):
            current_stage_flights.append(f_sorted[i])
            break

    # step 3 : listing current stage flights (not in prevoius stages and depart before the arrival of critical flight)
    
    for i in range(len(f_sorted)):
        if(f_sorted[i].included_in_stage==0 and f_sorted[i].planned_departure < current_stage_flights[0].planned_arrival):
            f_sorted[i].included_in_stage = current_stage
            current_stage_flights.append(f_sorted[i])
            

    for row in range(2, ws.max_row):
        for i in range(len(current_stage_flights)):
            if (ws[row][0].value == current_stage_flights[i].flight_id):
                ws[row][4].value = current_stage

    wb.save("flights.xlsx")

    return current_stage_flights




'''for row in range(len(current_stage_flights)):
        res.cell(row=row+2,column = 1).value = current_stage_flights[row].flight_id
        res.cell(row=row+2,column = 2).value = current_stage_flights[row].planned_departure
        res.cell(row=row+2,column = 3).value = current_stage_flights[row].planned_arrival
        res.cell(row=row+2,column = 4).value = current_stage_flights[row].delay
        res.cell(row=row+2,column = 5).value = current_stage_flights[row].included_in_stage 


'''
